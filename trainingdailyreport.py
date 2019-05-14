#!/usr/bin/env python
# coding: utf-8

# In[40]:


#openpyxlをインポート、workbookを作成、python.xlsxで保存　wb=workbook,ws=worksheet
import openpyxl as px
wb = px.Workbook()
wb.save('研修日報週報.xlsx')


# In[41]:


#ws名をテンプレートに変更
ws=wb.active
ws.title="Sheet1"


# In[42]:


#セルの高さを20.1に変更 
for i in range(100):
    ws.row_dimensions[i].height=20.1
wb.save('研修日報週報.xlsx')


# In[43]:


#セルの幅を3.94に変更　
for c in range(ord('A'),ord('Z')+1):
    ws.column_dimensions[chr(c)].width = 3.94
wb.save('研修日報週報.xlsx')


# In[44]:


#いちばんうえの「研修日報週報」を作成、フォント,サイズを変更、太字、中央揃え
ws.merge_cells("A1:Y2")
from openpyxl.styles.fonts import Font
ws["A1"].font = Font(
    name = '游ゴシック',
    size = 20,
    b = True
)
ws["A1"] = "研修日報週報"
from openpyxl.styles.alignment import Alignment
ws["A1"].alignment = Alignment(
    horizontal = 'center',
    vertical = 'center'
)
wb.save('研修日報週報.xlsx')


# In[45]:


#対象期間　
ws.merge_cells("B5:C5")


#ws["B5"].border = bottom

ws["B5"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["B5"] = "対象期間"
ws["B5"].alignment = Alignment(
    vertical = 'center'
)

ws.merge_cells("D5:G5")
ws["D5"].font = Font(
    name = '游ゴシック',
    size = 11
)

ws['D5']= '2019/04/15'
ws["D5"].alignment = Alignment(
    vertical = 'center',
    horizontal = 'right'
)

ws["H5"] = "～"
ws["H5"].alignment = Alignment(
    vertical = 'center',
    horizontal = 'center'
)

ws.merge_cells("I5:K5")
ws["I5"].font = Font(
    name = '游ゴシック',
    size = 11
)

ws['I5']= '=D5+4'
ws["I5"].alignment = Alignment(
    vertical = 'center',
    horizontal = 'left'
)

wb.save('研修日報週報.xlsx')


# In[46]:


#所属　
ws.merge_cells("B6:C6")
ws.merge_cells("D6:K6")

#ws["B6"].border = bottom
#ws["D6"].border = bottom

ws["B6"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["B6"] = "所属"
ws["B6"].alignment = Alignment(
    vertical = 'center',
    horizontal = 'left'
)

ws["D6"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["D6"] = "アンドロボティクス株式会社"
ws["D6"].alignment = Alignment(
    vertical = 'center',
    horizontal = 'center'
)

wb.save('研修日報週報.xlsx')


# In[47]:


#氏名　
ws.merge_cells("B7:C7")
ws.merge_cells("D7:K7")

#ws["B6"].border = bottom
#ws["D6"].border = bottom

ws["B7"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["B7"] = "氏名"
ws["B7"].alignment = Alignment(
    vertical = 'center',
    horizontal = 'left'
)

ws["D7"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["D7"] = "佐藤太郎"
ws["D7"].alignment = Alignment(
    vertical = 'center',
    horizontal = 'center'
)

wb.save('研修日報週報.xlsx')


# In[48]:


#管理～講師　(罫線)
ws.merge_cells("P4:R4")
ws.merge_cells("S4:U4")
ws.merge_cells("V4:X4")
ws.merge_cells("P5:R7")
ws.merge_cells("S5:U7")
ws.merge_cells("V5:X7")

ws["P4"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["P4"] = "管理"
ws["P4"].alignment = Alignment(
    vertical = 'center',
    horizontal = 'center'
)

ws["S4"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["S4"] = "自社管理"
ws["S4"].alignment = Alignment(
    vertical = 'center',
    horizontal = 'center'
)

ws["V4"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["V4"] = "講師"
ws["V4"].alignment = Alignment(
    vertical = 'center',
    horizontal = 'center'
)

wb.save('研修日報週報.xlsx')


# In[49]:


#今週のゴール
ws.merge_cells("B9:X9")
ws.merge_cells("B10:X13")

ws["B9"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["B9"] = "今週のゴール"
ws["B9"].alignment = Alignment(
    vertical = 'center',
    horizontal = 'center'
)

wb.save('研修日報週報.xlsx')


# In[50]:



ws.merge_cells("B51:X51")  #セル結合

ws["B51"].value = "備考"
ws["B51"].font = Font(
    name = '游ゴシック',
    size = 11
)


ws.merge_cells("B52:X53") #セル結合
ws.merge_cells("B46:X46")
ws.merge_cells("B47:X49")
ws.merge_cells("C40:H45")
ws.merge_cells("I40:P45")
ws.merge_cells("Q40:X45")
ws.merge_cells("C34:H39")
ws.merge_cells("I34:P39")
ws.merge_cells("Q34:X39")
ws.merge_cells("C22:H27")
ws.merge_cells("I22:P27")
ws.merge_cells("Q22:X27")
ws.merge_cells("C16:H21")
ws.merge_cells("I16:P21")
ws.merge_cells("Q16:X21")
ws.merge_cells("C15:H15")
ws.merge_cells("I15:P15")
ws.merge_cells("Q15:X15")
ws.merge_cells("C28:H33")
ws.merge_cells("I28:P33")
ws.merge_cells("Q28:X33")



wb.save("研修日報週報.xlsx")


# In[ ]:





# In[51]:

from openpyxl.styles.alignment import Alignment
for row in ws:
    for cell in row:
            cell.alignment = Alignment(horizontal = 'center', 
                                        vertical = 'center')
#シート全体を中央ぞろえ

ws["B15"].value = "日付"
ws["B15"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["C15"].value = "今日のゴール"
ws["C15"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["I15"].value = "どこまで出来たか"
ws["I15"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["Q15"].value = "問題点・懸念点など"
ws["Q15"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["B46"].value = "今週の振り返り・来週の目標"
ws["B46"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["B19"].value = "(月)"
ws["B19"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["B25"].value = "(火)"
ws["B25"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["B31"].value = "(水)"
ws["B31"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["B37"].value = "(木)"
ws["B37"].font = Font(
    name = '游ゴシック',
    size = 11
)
ws["B43"].value = "(金)"
ws["B43"].font = Font(
    name = '游ゴシック',
    size = 11
)



wb.save("研修日報週報.xlsx")


# In[61]:


from openpyxl.styles.borders import Border, Side
border = Border(top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')            
                )

#B5K7
bottomborder = Border(bottom=Side(style='thin', color='000000')
                     )                      
for row_num in range(5,8):
    for col_num in range(2,12):
        ws.cell(row=row_num ,column=col_num).border = bottomborder

        
leftborder = Border(left=Side(style='thin', color='000000')
                     )                      
for row_num in range(15,50):
    for col_num in range(2,3):
        ws.cell(row=row_num ,column=col_num).border = leftborder
        
lborder = Border(bottom=Side(style='thin', color='000000'),
                 left=Side(style='thin', color='000000')
                     )                      
ws["B21"].border = lborder
ws["B27"].border = lborder
ws["B33"].border = lborder
ws["B39"].border = lborder
ws["B45"].border = lborder
ws["B46"].border = lborder
ws["B49"].border = lborder
        
#B15X50
for row_num in range(15,50):
    for col_num in range(3,25):
        ws.cell(row=row_num ,column=col_num).border = border

#P4X7
for row_num in range(4,8):
    for col_num in range(16,25):
        ws.cell(row=row_num ,column=col_num).border = border
        
#B9X13
for row_num in range(9,14):
    for col_num in range(2,25):
        ws.cell(row=row_num ,column=col_num).border = border
        
#B51X53
for row_num in range(51,54):
    for col_num in range(2,25):
        ws.cell(row=row_num ,column=col_num).border = border

ws["B15"].border = border
        
wb.save("研修日報週報.xlsx")






